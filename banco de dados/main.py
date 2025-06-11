import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def criar_dataframe():
    """Cria um DataFrame com os dados fornecidos"""
    dados = {
        'Nome': [
            'Meu Troco', 'Prime', 'Canteiro', 'João de Barro', 'Telsonn', 'Bem vivido', 'Skeed', 'School King', 'AMSA',
            'Energy Greed', 'Guiar', 'Healpp tech', 'MertricDev', 'Higia', 'Printese', 'SmartGears', 'Carreira Hub',
            'Inclusão na Prática', 'Innovacci Alimentos Inteligentes', 'Lis Saúde', 'MAKE A VISION', 'Ponto a Ponto',
            'Reversa com Sistema Solar Offgrid', 'Shopping Cidadão', 'Sistema de Gerenciamento de Fichas Online - SisGeFiO',
            'Smart Solutions IoT', 'Vexa Lab', 'Alimentos Upcycled', 'Cata cata', 'Chemall', 'Coco Dog', 'Ecocamping Lumiar',
            'Ookami', 'Universo das plantas', 'BioPoliTech', 'ForCE Metabolomics', "K'auy bebidas", 'Mãe do Mato Foodtech',
            'Qualileite Neonatal', 'Sitio Mangara', 'Smart Chef'
        ],
        'Ano': [
            2020, 2020, 2020, 2020, 2020, 2020, 2020, 2020, 2020,
            2021, 2021, 2021, 2021, 2021, 2021, 2021, 2021,
            2022, 2022, 2022, 2022, 2022, 2022, 2022, 2022, 2022, 2022,
            2023, 2023, 2023, 2023, 2023, 2023, 2023,
            2024, 2024, 2024, 2024, 2024, 2024, 2024
        ]
    }
    return pd.DataFrame(dados)

def agrupar_incubadas_por_ano(df):
    """Agrupa as incubadas por ano de graduação"""
    # Agrupa por ano e coleta os nomes em uma lista
    grouped = df.groupby('Ano')['Nome'].apply(list).reset_index()
    
    # Conta o número de incubadas por ano
    grouped['Quantidade'] = grouped['Nome'].apply(len)
    
    return grouped

def salvar_excel_formatado(df, grouped_df, nome_arquivo='incubadas_por_ano.xlsx'):
    """Salva os resultados em um arquivo Excel com formatação"""
    
    # Cria um novo workbook
    wb = Workbook()
    
    # Remove a sheet padrão criada automaticamente
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 1. Sheet: Resumo por Ano
    ws1 = wb.create_sheet(title="Resumo por Ano")
    
    # Cabeçalho
    ws1.append(['Ano', 'Quantidade de Incubadas', 'Lista de Incubadas'])
    
    # Preenche os dados
    for _, row in grouped_df.iterrows():
        ws1.append([
            row['Ano'],
            row['Quantidade'],
            ', '.join(row['Nome'])
        ])
    
    # Formatação da sheet Resumo
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Ajusta largura das colunas
    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 80
    
    # 2. Sheet: Lista Completa
    ws2 = wb.create_sheet(title="Lista Completa")
    
    # Cabeçalho
    ws2.append(['Nome da Incubada', 'Ano de Graduação'])
    
    # Preenche os dados
    for r in dataframe_to_rows(df, index=False, header=False):
        ws2.append(r)
    
    # Formatação da sheet Lista Completa
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Ajusta largura das colunas
    ws2.column_dimensions['A'].width = 50
    ws2.column_dimensions['B'].width = 20
    
    # 3. Sheet: Dados para Dashboard (formato tabela)
    ws3 = wb.create_sheet(title="Dados Dashboard")
    
    # Cria uma versão expandida dos dados (uma linha por incubada)
    dados_dashboard = []
    for _, row in grouped_df.iterrows():
        for nome in row['Nome']:
            dados_dashboard.append([nome, row['Ano']])
    
    # Adiciona cabeçalho
    ws3.append(['Nome da Incubada', 'Ano de Graduação'])
    
    # Adiciona dados
    for linha in dados_dashboard:
        ws3.append(linha)
    
    # Formatação
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Ajusta largura
    ws3.column_dimensions['A'].width = 50
    ws3.column_dimensions['B'].width = 20
    
    # Salva o arquivo
    wb.save(nome_arquivo)

def main():
    try:
        # Cria o dataframe com os dados
        df = criar_dataframe()
        
        # Agrupa as incubadas por ano
        grouped = agrupar_incubadas_por_ano(df)
        
        # Exibe os resultados no console
        print("\nLISTA DE INCUBADAS GRADUADAS POR ANO")
        print("===================================")
        print(f"Total de anos analisados: {len(grouped)}")
        print(f"Total de incubadas: {len(df)}")
        
        # Salva os resultados em um arquivo Excel formatado
        salvar_excel_formatado(df, grouped)
        print(f"\nResultados salvos em 'incubadas_por_ano.xlsx'")
        print("Planilhas criadas:")
        print("- Resumo por Ano")
        print("- Lista Completa")
        print("- Dados Dashboard (para uso em gráficos)")
        
    except Exception as e:
        print(f"\nOcorreu um erro: {e}")

if __name__ == "__main__":
    main()